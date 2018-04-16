# -*- coding=utf-8 -*-
import openpyxl.styles

class runner(object):
    def __init__(self):
        self._excel_name="KBSS_统一账户系统对接系统关系清单V1.1.xlsx"

    def process(self):
        wb=openpyxl.load_workbook(self._excel_name,read_only=True)
        first_sheet=wb.worksheets[1]
        map_lbm={}
        for row in first_sheet:
            if row[1].value!="empty":
                map_lbm[row[1].value]=row[1].value
                #print(row[1].value)

        second_sheet=wb.worksheets[2]
        fil={}
        seq=0
        wbr=openpyxl.Workbook()
        wbr.remove(wbr.active)
        sheet1=wbr.create_sheet("kbss2wfis全量接口")

        sheet2=wbr.create_sheet("kbss2wfis差异接口")
        sheet2.append(["所属系统","接口LBM编号","中转LBM","账户LBM编号","接口描述"])

        for row in second_sheet:
            val=row[1].value
            sheet1.append([row[0].value,row[1].value,row[2].value,row[3].value,row[4].value])
            if (val!="empty") and(not (val in map_lbm))and (not (val in fil)):
                seq+=1
                print(seq,val,row[2].value,row[3].value,row[4].value)
                fil[val]=val
                sheet2.append([row[0].value,row[1].value,row[2].value,row[3].value,row[4].value])
        wb.close()
        wbr.save("kbss2wfisV1.0.xlsx")
        wbr.close()

if __name__=="__main__":
    print("starting...")
    r=runner()
    r.process()
    print("exted")
